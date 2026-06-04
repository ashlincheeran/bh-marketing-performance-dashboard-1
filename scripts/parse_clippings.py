#!/usr/bin/env python3
"""
Parse the Cencorp / betterhomes PR press-clipping workbook into clean,
normalized datasets used to seed the dashboard + Supabase.

Inputs : data/source/Cencorp_PR_Press_clipping.xlsx
Outputs: data/mentions.json   (one row per press clip, normalized + deduped)
         data/outlets.json    (reference table: outlet -> tier / eav / reach)
         data/mentions.csv     (same as mentions.json, for Supabase CSV import)
         data/outlets.csv      (same as outlets.json, for Supabase CSV import)
         data/import_summary.json

Run:  python3 scripts/parse_clippings.py
"""
from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import statistics
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(ROOT, "data", "source", "Cencorp_PR_Press_clipping.xlsx")
OUT_DIR = os.path.join(ROOT, "data")

# Per-year clip sheets to import. "Leasing Clips" has no header row.
YEAR_SHEETS = ["2020", "2021", "2022", "2023", "2024", "2025", "2026"]
EXTRA_SHEETS = ["Leasing Clips"]

DATE_HEADERS = {"date", "cc", "ccc"}  # header text (lower) that marks the date col


def norm_tier(raw: str | None) -> str:
    t = (raw or "").strip().lower()
    if t in ("tier 1", "tier 1 international", "tier 1 intl"):
        return "T1-Global"
    if t == "tier 1 local":
        return "T1-Local"
    if t.startswith("tier 2"):
        return "T2"
    if t == "tier 3":
        return "T3"
    return "Other"


# Sub-brand resolution. Order matters: more specific brands win over the
# generic "betterhomes" so sub-brand activity stays visible in combos
# like "Betterhomes/CRC".
BRAND_RULES = [
    ("CRC", lambda s: "crc" in s),
    ("PRIME", lambda s: "prime" in s),
    ("Off-plan", lambda s: "off" in s and "plan" in s),
    ("Lomond", lambda s: "lomond" in s),
    ("BetterStay", lambda s: "betterstay" in s or "better stay" in s),
    ("BH Mortgages", lambda s: "mortgage" in s),
    ("Top 50 Homes", lambda s: "top 50" in s),
    ("Linda's", lambda s: "linda" in s),
    ("Cencorp", lambda s: "cencorp" in s),
]


def norm_brand(raw: str | None) -> str:
    s = (raw or "").strip().lower()
    if not s or s == "-":
        return "betterhomes"
    for name, test in BRAND_RULES:
        if test(s):
            return name
    if "better" in s or s in ("bh", "str bh", "pm betterhomes", "openhub"):
        return "betterhomes"
    return "betterhomes"


def to_iso_date(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else None


def to_num(v):
    if v is None or v == "":
        return None
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    return int(round(n))


def clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s or None


def first_url(*vals) -> str | None:
    for v in vals:
        s = clean(v)
        if s and (s.lower().startswith("http") or s.lower() == "print"):
            return s
    # fall back to any non-#REF cell
    for v in vals:
        s = clean(v)
        if s and s != "#REF!":
            return s
    return None


COUNTRY_RE = re.compile(r"\(([^)]+)\)\s*$")


def country_from_outlet(name: str | None) -> str | None:
    """Pull a country out of names like 'Super Guide Veronica (Netherlands)'."""
    if not name:
        return None
    m = COUNTRY_RE.search(name)
    if m:
        c = m.group(1).strip()
        if re.fullmatch(r"[A-Za-z .'\-]{3,}", c) and c.lower() != "print":
            return c
    return None


def derive_media_type(url: str | None, title: str | None) -> str:
    u = (url or "").lower()
    t = (title or "").lower()
    if u == "print":
        return "print"
    if "podcast" in t:
        return "podcast"
    if u.startswith("http"):
        return "online"
    return "other"


def derive_tags(title: str | None, sheet: str) -> list[str]:
    """Light topic tagging — extend freely; the column is a flexible array."""
    tags: set[str] = set()
    if sheet == "Leasing Clips":
        tags.add("leasing")
    t = (title or "").lower()
    if any(k in t for k in ("rent", "tenant", "lease", "leasing")):
        tags.add("leasing")
    if "report" in t:
        tags.add("market-report")
    if any(k in t for k in ("ceo", "appoint", "steps down", "director", "leadership")):
        tags.add("leadership")
    if ("off-plan" in t) or ("off plan" in t) or ("offplan" in t):
        tags.add("off-plan")
    if "top 50" in t:
        tags.add("top-50")
    if "ramadan" in t:
        tags.add("ramadan")
    return sorted(tags)


def header_map(ws):
    """Map normalized header text -> column index (1-based) from row 1."""
    hm = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            hm[str(v).strip().lower()] = c
    return hm


def find_col(hm, *names):
    for n in names:
        if n in hm:
            return hm[n]
    return None


def parse_sheet(ws, name: str):
    rows = []
    if name == "Leasing Clips":  # no header: Date, Tier, Media, Title, Link, Brand
        cols = {"date": 1, "tier": 2, "media": 3, "title": 4, "link": 5,
                "link2": None, "brand": 6, "cost": None, "reach": None}
        start = 1
    else:
        hm = header_map(ws)
        date_col = next((hm[h] for h in hm if h in DATE_HEADERS), 1)
        # the 2nd "link" column (some sheets have two)
        link_cols = [c for h, c in hm.items() if h == "link"]
        cols = {
            "date": date_col,
            "tier": find_col(hm, "media tier", "tier"),
            "media": find_col(hm, "media", "outlet", "publication"),
            "cost": find_col(hm, "cost", "eav"),
            "reach": find_col(hm, "reach"),
            "title": find_col(hm, "title", "headline"),
            "link": link_cols[0] if link_cols else find_col(hm, "link", "url"),
            "link2": link_cols[1] if len(link_cols) > 1 else None,
            "brand": find_col(hm, "brand"),
        }
        start = 2

    def cell(r, key):
        c = cols.get(key)
        return ws.cell(row=r, column=c).value if c else None

    for r in range(start, ws.max_row + 1):
        outlet = clean(cell(r, "media"))
        title = clean(cell(r, "title"))
        if not outlet and not title:
            continue
        date = to_iso_date(cell(r, "date"))
        link = first_url(cell(r, "link"), cell(r, "link2"))
        rows.append({
            "date": date,
            "year": int(date[:4]) if date else None,
            "month": int(date[5:7]) if date else None,
            "tier": norm_tier(cell(r, "tier")),
            "tier_raw": clean(cell(r, "tier")),
            "outlet": outlet,
            "title": title,
            "url": link,
            "eav": to_num(cell(r, "cost")),
            "reach": to_num(cell(r, "reach")),
            "brand": norm_brand(cell(r, "brand")),
            "brand_raw": clean(cell(r, "brand")),
            "sentiment": None,            # not in source — enriched later
            "media_type": derive_media_type(link, title),
            "language": None,             # enriched later
            "tags": derive_tags(title, name),
            "source": "historical_import",
            "metadata": {"sheet": name},
            "sheet": name,
        })
    return rows


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    raw = []
    per_sheet = {}
    for name in YEAR_SHEETS + EXTRA_SHEETS:
        if name not in wb.sheetnames:
            continue
        rows = parse_sheet(wb[name], name)
        per_sheet[name] = len(rows)
        raw.extend(rows)

    # Dedup by (date, outlet, title). Keep the richest record (most eav/reach).
    best: dict[str, dict] = {}
    for m in raw:
        key = "|".join([
            m["date"] or "",
            (m["outlet"] or "").lower(),
            (m["title"] or "").lower()[:120],
        ])
        m["id"] = hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]
        cur = best.get(key)
        if cur is None:
            best[key] = m
            continue
        score = (m["eav"] is not None) + (m["reach"] is not None) + (m["url"] is not None)
        cscore = (cur["eav"] is not None) + (cur["reach"] is not None) + (cur["url"] is not None)
        if score > cscore:
            best[key] = m

    mentions = list(best.values())
    mentions.sort(key=lambda m: (m["date"] or "0000-00-00"), reverse=True)

    # Build outlet reference table.
    by_outlet = defaultdict(list)
    for m in mentions:
        if m["outlet"]:
            by_outlet[m["outlet"]].append(m)

    outlets = []
    for name, clips in sorted(by_outlet.items()):
        tiers = Counter(c["tier"] for c in clips if c["tier"] != "Other")
        tier = tiers.most_common(1)[0][0] if tiers else "Other"
        eavs = [c["eav"] for c in clips if c["eav"] is not None]
        reaches = [c["reach"] for c in clips if c["reach"] is not None]
        dates = [c["date"] for c in clips if c["date"]]
        outlets.append({
            "outlet": name,
            "tier": tier,
            "country": country_from_outlet(name),
            "language": None,
            "default_eav": int(statistics.median(eavs)) if eavs else None,
            "default_reach": int(statistics.median(reaches)) if reaches else None,
            "clip_count": len(clips),
            "first_seen": min(dates) if dates else None,
            "last_seen": max(dates) if dates else None,
        })
    outlets.sort(key=lambda o: o["clip_count"], reverse=True)

    # ---- write JSON ----
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "mentions.json"), "w", encoding="utf-8") as f:
        json.dump(mentions, f, ensure_ascii=False, indent=1)
    with open(os.path.join(OUT_DIR, "outlets.json"), "w", encoding="utf-8") as f:
        json.dump(outlets, f, ensure_ascii=False, indent=1)

    # ---- write CSV (for Supabase import) ----
    mention_cols = ["id", "date", "year", "month", "tier", "tier_raw", "outlet",
                    "title", "url", "eav", "reach", "brand", "brand_raw",
                    "sentiment", "media_type", "language", "source"]
    with open(os.path.join(OUT_DIR, "mentions.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=mention_cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(mentions)
    outlet_cols = ["outlet", "tier", "country", "language", "default_eav",
                   "default_reach", "clip_count", "first_seen", "last_seen"]
    with open(os.path.join(OUT_DIR, "outlets.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=outlet_cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(outlets)

    # ---- summary ----
    by_year = Counter(m["year"] for m in mentions if m["year"])
    by_tier = Counter(m["tier"] for m in mentions)
    by_brand = Counter(m["brand"] for m in mentions)
    summary = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "source_file": "data/source/Cencorp_PR_Press_clipping.xlsx",
        "rows_read_per_sheet": per_sheet,
        "total_raw_rows": len(raw),
        "total_after_dedup": len(mentions),
        "distinct_outlets": len(outlets),
        "by_year": dict(sorted(by_year.items())),
        "by_tier": dict(by_tier.most_common()),
        "by_brand": dict(by_brand.most_common()),
        "with_eav": sum(1 for m in mentions if m["eav"] is not None),
        "with_reach": sum(1 for m in mentions if m["reach"] is not None),
        "with_url": sum(1 for m in mentions if m["url"]),
    }
    with open(os.path.join(OUT_DIR, "import_summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
